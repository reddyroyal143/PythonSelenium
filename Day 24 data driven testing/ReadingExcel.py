import openpyxl

file="C:\\Users\\royal\\Desktop\\Selenium\\Book1.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row
cols=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='    ')
    print()




