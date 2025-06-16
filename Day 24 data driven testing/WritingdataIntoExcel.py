import openpyxl

#same data
file="C:\\Users\\royal\\Desktop\\Selenium\\Book2.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]  # or sheet=workbook.active if u have only i sheet

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="welcome"
workbook.save(file)

#multiple data
file="C:\\Users\\royal\\Desktop\\Selenium\\Book2.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]  # or sheet=workbook.active if u have only i sheet

sheet.cell(1,1).value=123
sheet.cell(1,2).value='smith'
sheet.cell(1,3).value='engineer'

sheet.cell(2,1).value=456
sheet.cell(2,2).value='john'
sheet.cell(2,3).value='manager'

sheet.cell(3,1).value=789
sheet.cell(3,2).value='david'
sheet.cell(3,3).value='developer'

workbook.save(file)



